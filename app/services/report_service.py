from typing import List, Optional
from uuid import UUID
from datetime import datetime, date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from app.models.attendance import Attendance, AttendanceStatus, AttendanceType
from app.models.employee import Employee
from app.crud.organization import organization as org_crud
from app.crud.employee import employee as employee_crud
from app.schemas.report import AttendanceReportRow, AttendanceReport


class ReportService:
    """Service de génération de rapports"""
    
    def generate_attendance_report(
        self,
        db: Session,
        organization_id: UUID,
        start_date: date,
        end_date: date,
        employee_ids: Optional[List[int]] = None,
        department: Optional[str] = None
    ) -> AttendanceReport:
        """
        Générer un rapport de présence
        """
        # Récupérer l'organisation
        org = org_crud.get(db, id=organization_id)
        
        # Construire la requête des employés
        query = db.query(Employee).filter(Employee.organization_id == organization_id)
        
        if employee_ids:
            query = query.filter(Employee.id.in_(employee_ids))
        
        if department:
            query = query.filter(Employee.department == department)
        
        employees = query.all()
        
        # Calculer les statistiques pour chaque employé
        rows = []
        total_days = (end_date - start_date).days + 1
        
        for emp in employees:
            # Compter les présences par statut
            present = self._count_attendances(
                db, emp.id, start_date, end_date, AttendanceStatus.PRESENT
            )
            absent = self._count_attendances(
                db, emp.id, start_date, end_date, AttendanceStatus.ABSENT
            )
            late = self._count_attendances(
                db, emp.id, start_date, end_date, AttendanceStatus.LATE
            )
            on_leave = self._count_attendances(
                db, emp.id, start_date, end_date, AttendanceStatus.ON_LEAVE
            )
            
            # Calculer les heures totales
            total_hours = self._calculate_total_hours(
                db, emp.id, start_date, end_date
            )
            
            # Taux de présence
            attendance_rate = (present / total_days * 100) if total_days > 0 else 0
            
            rows.append(AttendanceReportRow(
                employee_id=emp.id,
                employee_name=f"{emp.first_name} {emp.last_name}",
                badge_id=emp.badge_id,
                department=emp.department,
                total_days=total_days,
                present_days=present,
                absent_days=absent,
                late_days=late,
                on_leave_days=on_leave,
                attendance_rate=round(attendance_rate, 2),
                total_hours=round(total_hours, 2)
            ))
        
        # Calculer le résumé
        summary = {
            "total_present": sum(r.present_days for r in rows),
            "total_absent": sum(r.absent_days for r in rows),
            "total_late": sum(r.late_days for r in rows),
            "total_on_leave": sum(r.on_leave_days for r in rows),
            "average_attendance_rate": round(
                sum(r.attendance_rate for r in rows) / len(rows), 2
            ) if rows else 0,
            "total_hours": round(sum(r.total_hours for r in rows), 2)
        }
        
        return AttendanceReport(
            organization_name=org.name,
            period=f"{start_date} au {end_date}",
            start_date=start_date,
            end_date=end_date,
            total_employees=len(employees),
            rows=rows,
            summary=summary
        )
    
    def _count_attendances(
        self,
        db: Session,
        employee_id: int,
        start_date: date,
        end_date: date,
        status: AttendanceStatus
    ) -> int:
        """Compter les pointages par statut"""
        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time())
        
        # Compter les jours uniques avec ce statut
        count = db.query(func.count(func.distinct(func.date(Attendance.timestamp)))).filter(
            Attendance.employee_id == employee_id,
            Attendance.status == status,
            Attendance.timestamp >= start_dt,
            Attendance.timestamp <= end_dt
        ).scalar()
        
        return count or 0
    
    def _calculate_total_hours(
        self,
        db: Session,
        employee_id: int,
        start_date: date,
        end_date: date
    ) -> float:
        """Calculer le total d'heures travaillées"""
        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time())
        
        # Récupérer tous les pointages
        attendances = db.query(Attendance).filter(
            Attendance.employee_id == employee_id,
            Attendance.timestamp >= start_dt,
            Attendance.timestamp <= end_dt
        ).order_by(Attendance.timestamp).all()
        
        total_hours = 0.0
        check_in_time = None
        
        for att in attendances:
            if att.attendance_type == AttendanceType.CHECK_IN:
                check_in_time = att.timestamp
            elif att.attendance_type == AttendanceType.CHECK_OUT and check_in_time:
                # Calculer la différence
                duration = (att.timestamp - check_in_time).total_seconds() / 3600
                total_hours += duration
                check_in_time = None
        
        return total_hours
    
    def export_to_csv(self, report: AttendanceReport) -> str:
        """Exporter le rapport en CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # En-têtes
        writer.writerow([
            "ID Employé", "Nom", "Badge", "Département",
            "Jours Total", "Présents", "Absents", "Retards",
            "Congés", "Taux Présence (%)", "Heures Total"
        ])
        
        # Données
        for row in report.rows:
            writer.writerow([
                row.employee_id,
                row.employee_name,
                row.badge_id,
                row.department or "N/A",
                row.total_days,
                row.present_days,
                row.absent_days,
                row.late_days,
                row.on_leave_days,
                row.attendance_rate,
                row.total_hours
            ])
        
        return output.getvalue()
    
    def export_to_excel(self, report: AttendanceReport) -> bytes:
        """Exporter le rapport en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport de Présence"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Titre
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = f"Rapport de Présence - {report.organization_name}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Période
        ws.merge_cells('A2:K2')
        period_cell = ws['A2']
        period_cell.value = f"Période: {report.period}"
        period_cell.alignment = Alignment(horizontal='center')
        
        # En-têtes
        headers = [
            "ID", "Nom Complet", "Badge", "Département",
            "Jours Total", "Présents", "Absents", "Retards",
            "Congés", "Taux (%)", "Heures"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row_idx, row_data in enumerate(report.rows, start=5):
            ws.cell(row=row_idx, column=1, value=row_data.employee_id)
            ws.cell(row=row_idx, column=2, value=row_data.employee_name)
            ws.cell(row=row_idx, column=3, value=row_data.badge_id)
            ws.cell(row=row_idx, column=4, value=row_data.department or "N/A")
            ws.cell(row=row_idx, column=5, value=row_data.total_days)
            ws.cell(row=row_idx, column=6, value=row_data.present_days)
            ws.cell(row=row_idx, column=7, value=row_data.absent_days)
            ws.cell(row=row_idx, column=8, value=row_data.late_days)
            ws.cell(row=row_idx, column=9, value=row_data.on_leave_days)
            ws.cell(row=row_idx, column=10, value=row_data.attendance_rate)
            ws.cell(row=row_idx, column=11, value=row_data.total_hours)
        
        # Résumé
        summary_row = len(report.rows) + 6
        ws.cell(row=summary_row, column=1, value="RÉSUMÉ").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value=f"Taux moyen: {report.summary['average_attendance_rate']}%")
        ws.cell(row=summary_row + 2, column=1, value=f"Total heures: {report.summary['total_hours']}h")
        
        # Ajuster les largeurs
        for col in range(1, 12):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_to_pdf(self, report: AttendanceReport) -> bytes:
        """Exporter le rapport en PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title = Paragraph(
            f"<b>Rapport de Présence - {report.organization_name}</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Période
        period = Paragraph(f"Période: {report.period}", styles['Normal'])
        elements.append(period)
        elements.append(Spacer(1, 20))
        
        # Tableau des données
        data = [
            ["Nom", "Badge", "Dept", "Présents", "Absents", "Retards", "Taux (%)"]
        ]
        
        for row in report.rows:
            data.append([
                row.employee_name,
                row.badge_id,
                row.department or "N/A",
                str(row.present_days),
                str(row.absent_days),
                str(row.late_days),
                f"{row.attendance_rate}%"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Résumé
        summary_text = f"""
        <b>RÉSUMÉ</b><br/>
        Taux de présence moyen: {report.summary['average_attendance_rate']}%<br/>
        Total heures travaillées: {report.summary['total_hours']}h<br/>
        Nombre total d'employés: {report.total_employees}
        """
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        
        doc.build(elements)
        buffer.seek(0)
        
        return buffer.getvalue()


# Instance globale
report_service = ReportService()